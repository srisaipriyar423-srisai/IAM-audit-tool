"""
🔐 IAM User Access Review & ITGC Evaluation Tool
Main Streamlit application.
"""

import io
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from data.users import USERS, ROLE_CATALOG, RBAC_MATRIX, SOD_RULES
from logic.risk_engine import calculate_risk
from logic.sod_engine import get_all_sod_violations

# ═══════════════════════════════════════════════════════════════════
# Page config & CSS
# ═══════════════════════════════════════════════════════════════════
st.set_page_config(page_title="IAM Audit Tool", page_icon="🔐", layout="wide")

CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@300;400;500;600;700&family=Inter:wght@300;400;500;600;700&display=swap');

/* ── Global ────────────────────────────────────────── */
html, body, [class*="css"] {
    font-family: 'Inter', 'JetBrains Mono', monospace;
}
.stApp {
    background: linear-gradient(135deg, #0a0c10 0%, #0d1117 50%, #101820 100%);
    color: #c9d1d9;
}

/* ── Sidebar ───────────────────────────────────────── */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d1117 0%, #161b22 100%);
    border-right: 1px solid #21262d;
}
section[data-testid="stSidebar"] .stMarkdown h1,
section[data-testid="stSidebar"] .stMarkdown h2,
section[data-testid="stSidebar"] .stMarkdown h3 {
    color: #58a6ff;
}

/* ── Metric cards ──────────────────────────────────── */
[data-testid="stMetric"] {
    background: linear-gradient(135deg, #161b22, #1c2333);
    border: 1px solid #30363d;
    border-radius: 12px;
    padding: 16px 20px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.3);
}
[data-testid="stMetric"] label {
    color: #8b949e !important;
    font-size: 0.8rem !important;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}
[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: #f0f6fc !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-weight: 700 !important;
}

/* ── Tabs ──────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {
    gap: 8px;
    background: #161b22;
    border-radius: 12px;
    padding: 6px;
    border: 1px solid #21262d;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    padding: 8px 20px;
    color: #8b949e;
    font-weight: 500;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #1f6feb, #388bfd) !important;
    color: #ffffff !important;
}

/* ── Expanders ─────────────────────────────────────── */
.streamlit-expanderHeader {
    background: #161b22;
    border: 1px solid #30363d;
    border-radius: 10px;
    color: #c9d1d9;
    font-weight: 600;
}

/* ── DataFrames ────────────────────────────────────── */
[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
}

/* ── Buttons ───────────────────────────────────────── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #238636, #2ea043) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    padding: 8px 24px !important;
    transition: all 0.3s ease;
}
.stDownloadButton > button:hover {
    box-shadow: 0 0 20px rgba(46, 160, 67, 0.4) !important;
    transform: translateY(-1px);
}

/* ── Badge helper classes (used via markdown) ──────── */
.badge {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.3px;
}
.badge-critical { background: #ff333355; color: #ff6666; border: 1px solid #ff3333; }
.badge-high     { background: #ff8c0044; color: #ffaa44; border: 1px solid #ff8c00; }
.badge-medium   { background: #ffd70033; color: #ffd700; border: 1px solid #ffd700; }
.badge-low      { background: #39ff1422; color: #39ff14; border: 1px solid #39ff14; }

/* ── Card containers ───────────────────────────────── */
.rec-card {
    background: linear-gradient(135deg, #161b22, #1c2333);
    border: 1px solid #30363d;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 12px;
    transition: border-color 0.3s;
}
.rec-card:hover { border-color: #58a6ff; }

/* ── Divider ───────────────────────────────────────── */
hr {
    border: none;
    border-top: 1px solid #21262d;
    margin: 1.5rem 0;
}

/* ── Scrollbar ─────────────────────────────────────── */
::-webkit-scrollbar { width: 8px; height: 8px; }
::-webkit-scrollbar-track { background: #0d1117; }
::-webkit-scrollbar-thumb { background: #30363d; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #484f58; }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
# Analyze all users
# ═══════════════════════════════════════════════════════════════════

@st.cache_data
def analyze_all_users():
    results = []
    for user in USERS:
        analysis = calculate_risk(user, RBAC_MATRIX, SOD_RULES, ROLE_CATALOG)
        merged = {**user, **analysis}
        results.append(merged)
    return results


analyzed_users = analyze_all_users()
sod_df = get_all_sod_violations(analyzed_users)

# ═══════════════════════════════════════════════════════════════════
# Header
# ═══════════════════════════════════════════════════════════════════

st.markdown("""
<h1 style='text-align:center; background: linear-gradient(90deg, #58a6ff, #bc8cff, #f778ba);
-webkit-background-clip: text; -webkit-text-fill-color: transparent;
font-family: Inter, sans-serif; font-weight: 800; font-size: 2.2rem; margin-bottom:0;'>
🔐 User Access Review & ITGC Evaluation
</h1>
<p style='text-align:center; color:#8b949e; font-size:0.95rem; margin-top:4px;'>
Audit Period: <strong>FY 2025</strong> &nbsp;|&nbsp; System: <strong>Enterprise IAM</strong> &nbsp;|&nbsp; Status: <span style="color:#ffd700;">DRAFT</span>
</p>
""", unsafe_allow_html=True)

st.markdown("---")

# ═══════════════════════════════════════════════════════════════════
# Sidebar
# ═══════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("## ⚙ Audit Filters")
    risk_filter = st.selectbox("Risk Level Filter", ["All", "Critical", "High", "Medium", "Low"])
    departments = sorted(set(u["department"] for u in USERS))
    dept_filter = st.multiselect("Department Filter", departments, default=departments)
    status_filter = st.selectbox("Employment Status", ["All", "Active", "Inactive", "Terminated"])
    mfa_only = st.checkbox("Show only MFA non-compliant users", value=False)
    excess_only = st.checkbox("Show only excess access users", value=False)

    st.markdown("---")
    st.metric("Total Users Reviewed", len(USERS))
    st.metric("Audit Date", "2025-07-01")

# ── Apply filters ────────────────────────────────────────────────
def apply_filters(users):
    filtered = users[:]
    if risk_filter != "All":
        filtered = [u for u in filtered if u["risk_level"] == risk_filter]
    if dept_filter:
        filtered = [u for u in filtered if u["department"] in dept_filter]
    if status_filter != "All":
        filtered = [u for u in filtered if u["employment_status"] == status_filter]
    if mfa_only:
        filtered = [u for u in filtered if not u["mfa_enabled"]]
    if excess_only:
        filtered = [u for u in filtered if u["excess_roles"]]
    return filtered


# ═══════════════════════════════════════════════════════════════════
# Helper: color-code risk
# ═══════════════════════════════════════════════════════════════════
RISK_COLORS = {"Critical": "#ff3333", "High": "#ff8c00", "Medium": "#ffd700", "Low": "#39ff14"}

def risk_badge(level):
    c = RISK_COLORS.get(level, "#8b949e")
    return f'<span class="badge badge-{level.lower()}">{level}</span>'

def sensitivity_badge(sens):
    mapping = {"Critical": "critical", "High": "high", "Medium": "medium", "Low": "low"}
    cls = mapping.get(sens, "low")
    return f'<span class="badge badge-{cls}">{sens}</span>'

# ═══════════════════════════════════════════════════════════════════
# Plotly theme helper
# ═══════════════════════════════════════════════════════════════════
def dark_plotly_layout(fig, title=""):
    fig.update_layout(
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        title=dict(text=title, font=dict(size=18, color="#c9d1d9", family="Inter")),
        font=dict(family="Inter, JetBrains Mono", color="#8b949e"),
        margin=dict(l=40, r=40, t=60, b=40),
        legend=dict(bgcolor="rgba(0,0,0,0)"),
    )
    return fig

# ═══════════════════════════════════════════════════════════════════
# Recommendations generator
# ═══════════════════════════════════════════════════════════════════
@st.cache_data
def generate_recommendations(users_data):
    recs = []

    # 1 – Terminated users with access
    termed = [u for u in users_data if u["department"] == "Terminated" or u["employment_status"] == "Terminated"]
    termed_with_access = [u for u in termed if u["assigned_roles"]]
    if termed_with_access:
        names = ", ".join(u["name"] for u in termed_with_access)
        recs.append({
            "id": "REC-001", "priority": "Critical", "timeline": "Immediate (24 hrs)",
            "finding": f"Terminated employee(s) ({names}) retain active system access with assigned roles.",
            "action": "Immediately revoke all access for terminated employees. Implement automated deprovisioning tied to HR termination workflow. Conduct forensic review of recent activity.",
            "owner": "IT Security / HR",
        })

    # 2 – MFA disabled
    no_mfa = [u for u in users_data if not u["mfa_enabled"] and u["employment_status"] == "Active"]
    if no_mfa:
        names = ", ".join(u["name"] for u in no_mfa)
        recs.append({
            "id": "REC-002", "priority": "Critical", "timeline": "Within 7 days",
            "finding": f"{len(no_mfa)} active user(s) ({names}) do not have Multi-Factor Authentication enabled.",
            "action": "Enforce MFA enrollment for all users immediately. Prioritise users with elevated or critical roles. Update IAM policy to mandate MFA for all accounts.",
            "owner": "IT Security",
        })

    # 3 – Critical SoD violations
    crit_sod = [u for u in users_data if any(c["severity"] == "Critical" for c in u["sod_conflicts"])]
    if crit_sod:
        names = ", ".join(u["name"] for u in crit_sod)
        recs.append({
            "id": "REC-003", "priority": "Critical", "timeline": "Within 14 days",
            "finding": f"Critical Segregation of Duties violations found for: {names}. Users can both create and approve financial transactions.",
            "action": "Immediately remove conflicting role from each user. Implement compensating controls (dual approval, audit logging). Review all transactions processed by affected users in the last 90 days.",
            "owner": "Finance Director / IT Security",
        })

    # 4 – Excessive privileges
    excess_users = [u for u in users_data if len(u["excess_roles"]) >= 2]
    if excess_users:
        recs.append({
            "id": "REC-004", "priority": "High", "timeline": "Within 30 days",
            "finding": f"{len(excess_users)} user(s) have 2 or more roles beyond their job-title baseline (RBAC violation).",
            "action": "Conduct per-user access review with department managers. Remove unneeded roles and document any justified exceptions with a risk acceptance form.",
            "owner": "Department Managers / IT Security",
        })

    # 5 – Password policy
    old_pwd = [u for u in users_data if u["password_age_days"] > 90 and u["employment_status"] == "Active"]
    if old_pwd:
        recs.append({
            "id": "REC-005", "priority": "High", "timeline": "Within 14 days",
            "finding": f"{len(old_pwd)} active user(s) have passwords older than the 90-day policy threshold.",
            "action": "Force password reset for affected users. Enforce maximum password age of 90 days in IAM policy. Send automated reminders 14 days before expiry.",
            "owner": "IT Operations",
        })

    # 6 – Dormant accounts
    dormant = [u for u in users_data if u["days_since_login"] > 90]
    if dormant:
        names = ", ".join(u["name"] for u in dormant)
        recs.append({
            "id": "REC-006", "priority": "High", "timeline": "Within 30 days",
            "finding": f"{len(dormant)} account(s) ({names}) have not been used in over 90 days.",
            "action": "Disable dormant accounts and notify account owners. Implement automated dormancy detection that disables accounts after 90 days of inactivity.",
            "owner": "IT Operations / HR",
        })

    # 7 – Cross-department access
    cross = [u for u in users_data
             if u["department"] not in ("IT", "Compliance", "Terminated")
             and any(r in u["assigned_roles"] for r in ["admin_panel", "server_admin", "db_admin"])]
    if cross:
        names = ", ".join(u["name"] for u in cross)
        recs.append({
            "id": "REC-007", "priority": "Medium", "timeline": "Within 30 days",
            "finding": f"Non-IT user(s) ({names}) hold IT administrative roles (admin_panel, server_admin, or db_admin).",
            "action": "Review business justification for cross-department admin access. Remove unjustified roles and document exceptions with risk owner sign-off.",
            "owner": "IT Director / Compliance",
        })

    # 8 – RBAC refresh
    recs.append({
        "id": "REC-008", "priority": "Medium", "timeline": "Within 60 days",
        "finding": "RBAC role matrix has not been reviewed since initial deployment. Several job titles show systematic excess access.",
        "action": "Conduct full RBAC matrix review with all department heads. Update role definitions to reflect current job functions. Implement quarterly access certification campaigns.",
        "owner": "CISO / Compliance Officer",
    })

    return recs


recommendations = generate_recommendations(analyzed_users)

# ═══════════════════════════════════════════════════════════════════
# Excel export helpers
# ═══════════════════════════════════════════════════════════════════
HEADER_FILL = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
CELL_FONT = Font(color="1A1A2E", size=10, name="Calibri")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def style_ws(ws, df):
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(df) + 2)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)


def build_full_report_excel():
    wb = Workbook()

    # Sheet 1 — User Access Review
    ws1 = wb.active
    ws1.title = "User Access Review"
    user_df = pd.DataFrame([
        {
            "User ID": u["id"], "Name": u["name"], "Department": u["department"],
            "Job Title": u["job_title"], "Assigned Roles": ", ".join(u["assigned_roles"]),
            "# Roles": len(u["assigned_roles"]), "Excess Roles": ", ".join(u["excess_roles"]),
            "Risk Score": u["risk_score"], "Risk Level": u["risk_level"],
            "MFA Enabled": "Yes" if u["mfa_enabled"] else "No",
            "Password Age": u["password_age_days"],
            "Last Login": u["last_login"], "Employment Status": u["employment_status"],
            "Compliance Status": u["compliance_status"],
            "Risk Flags": " | ".join(u["risk_flags"]),
        }
        for u in analyzed_users
    ])
    for r_idx, row in enumerate(
        [user_df.columns.tolist()] + user_df.values.tolist(), start=1
    ):
        for c_idx, val in enumerate(row, start=1):
            ws1.cell(row=r_idx, column=c_idx, value=val)
    style_ws(ws1, user_df)

    # Sheet 2 — SoD Violations
    ws2 = wb.create_sheet("SoD Violations")
    for r_idx, row in enumerate(
        [sod_df.columns.tolist()] + sod_df.values.tolist(), start=1
    ):
        for c_idx, val in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=val)
    style_ws(ws2, sod_df)

    # Sheet 3 — Recommendations
    ws3 = wb.create_sheet("Recommendations")
    rec_df = pd.DataFrame([
        {
            "Rec ID": r["id"], "Priority": r["priority"],
            "Finding": r["finding"], "Action": r["action"],
            "Owner": r["owner"], "Timeline": r["timeline"],
            "Status": "Open",
        }
        for r in recommendations
    ])
    for r_idx, row in enumerate(
        [rec_df.columns.tolist()] + rec_df.values.tolist(), start=1
    ):
        for c_idx, val in enumerate(row, start=1):
            ws3.cell(row=r_idx, column=c_idx, value=val)
    style_ws(ws3, rec_df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sidebar export ───────────────────────────────────────────────
with st.sidebar:
    st.markdown("---")
    st.download_button(
        label="📥 Export Full Audit Report (Excel)",
        data=build_full_report_excel(),
        file_name="IAM_Audit_Report_FY2025.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ═══════════════════════════════════════════════════════════════════
# TABS
# ═══════════════════════════════════════════════════════════════════

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Overview",
    "👥 User Access Review",
    "🗂 Role Mapping (RBAC)",
    "⛔ SoD Conflict Analysis",
    "✅ Recommendations",
])

# ═══════════════════════════════════════════════════════════════════
# TAB 1 — Overview
# ═══════════════════════════════════════════════════════════════════
with tab1:
    st.markdown("### 📊 Executive Dashboard")

    # Counts
    risk_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for u in analyzed_users:
        risk_counts[u["risk_level"]] += 1

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Users", len(analyzed_users))
    c2.metric("🔴 Critical Risk", risk_counts["Critical"])
    c3.metric("🟠 High Risk", risk_counts["High"])
    c4.metric("🟡 Medium Risk", risk_counts["Medium"])
    c5.metric("🟢 Low Risk", risk_counts["Low"])

    st.markdown("---")

    # ── Risk distribution pie ─────────────────────────────────────
    col_left, col_right = st.columns(2)

    with col_left:
        pie_df = pd.DataFrame([
            {"Level": k, "Count": v} for k, v in risk_counts.items() if v > 0
        ])
        fig_pie = px.pie(
            pie_df, names="Level", values="Count",
            color="Level",
            color_discrete_map=RISK_COLORS,
            hole=0.45,
        )
        fig_pie.update_traces(
            textinfo="label+value+percent",
            textfont_size=13,
            marker=dict(line=dict(color="#0a0c10", width=2)),
        )
        dark_plotly_layout(fig_pie, "Risk Distribution")
        st.plotly_chart(fig_pie, use_container_width=True)

    # ── Users by department ───────────────────────────────────────
    with col_right:
        dept_data = {}
        for u in analyzed_users:
            dept_data.setdefault(u["department"], []).append(u["risk_score"])
        dept_df = pd.DataFrame([
            {"Department": d, "Users": len(scores), "Avg Risk": round(sum(scores) / len(scores), 1)}
            for d, scores in dept_data.items()
        ]).sort_values("Avg Risk", ascending=False)

        fig_dept = px.bar(
            dept_df, x="Department", y="Users",
            color="Avg Risk",
            color_continuous_scale=["#39ff14", "#ffd700", "#ff8c00", "#ff3333"],
            text="Users",
        )
        fig_dept.update_traces(textposition="outside")
        dark_plotly_layout(fig_dept, "Users by Department (colored by Avg Risk Score)")
        st.plotly_chart(fig_dept, use_container_width=True)

    # ── Top risk flags frequency ──────────────────────────────────
    flag_counter: dict[str, int] = {}
    for u in analyzed_users:
        for f in u["risk_flags"]:
            # Simplify the label
            key = f.split(":")[0] if ":" in f else f
            if key.startswith("User has"):
                key = "Excess Roles"
            elif key.startswith("Password age"):
                key = "Password > 90 days"
            elif key.startswith("Account dormant"):
                key = "Dormant Account (>90 days)"
            elif key.startswith("Holds"):
                key = "High Critical-Role Concentration"
            flag_counter[key] = flag_counter.get(key, 0) + 1

    flag_df = pd.DataFrame([
        {"Flag": k, "Users Affected": v}
        for k, v in sorted(flag_counter.items(), key=lambda x: -x[1])
    ])
    fig_flags = px.bar(
        flag_df, x="Users Affected", y="Flag", orientation="h",
        color="Users Affected",
        color_continuous_scale=["#58a6ff", "#bc8cff", "#ff6666"],
        text="Users Affected",
    )
    fig_flags.update_traces(textposition="outside")
    dark_plotly_layout(fig_flags, "Top Risk Flags — Frequency Across Users")
    fig_flags.update_layout(height=400)
    st.plotly_chart(fig_flags, use_container_width=True)

    # ── Top 5 riskiest users ──────────────────────────────────────
    st.markdown("#### 🔥 Top 5 Highest-Risk Users")
    top5 = sorted(analyzed_users, key=lambda u: -u["risk_score"])[:5]
    top5_df = pd.DataFrame([
        {
            "Name": u["name"],
            "Dept": u["department"],
            "Risk Score": u["risk_score"],
            "Risk Level": u["risk_level"],
            "# Flags": len(u["risk_flags"]),
        }
        for u in top5
    ])

    def color_risk_level(val):
        color = RISK_COLORS.get(val, "#8b949e")
        return f"color: {color}; font-weight: 700;"

    styled = top5_df.style.map(color_risk_level, subset=["Risk Level"])
    st.dataframe(styled, use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════════
# TAB 2 — User Access Review
# ═══════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("### 👥 User Access Review")
    filtered = apply_filters(analyzed_users)

    if not filtered:
        st.warning("No users match the current filters.")
    else:
        table_df = pd.DataFrame([
            {
                "User ID": u["id"],
                "Name": u["name"],
                "Department": u["department"],
                "Job Title": u["job_title"],
                "# Roles": len(u["assigned_roles"]),
                "Risk Score": u["risk_score"],
                "Risk Level": u["risk_level"],
                "MFA": "✅" if u["mfa_enabled"] else "❌",
                "Pwd Age": u["password_age_days"],
                "Compliance": u["compliance_status"],
            }
            for u in filtered
        ])

        def color_table_risk(val):
            color = RISK_COLORS.get(val, "#8b949e")
            return f"color: {color}; font-weight: 700;"

        styled_table = table_df.style.map(color_table_risk, subset=["Risk Level"])
        st.dataframe(styled_table, use_container_width=True, hide_index=True, height=460)

        # ── Detail selector ───────────────────────────────────────
        st.markdown("---")
        user_names = [u["name"] for u in filtered]
        selected_name = st.selectbox("🔎 Select a user for detailed review", user_names)
        sel_user = next(u for u in filtered if u["name"] == selected_name)

        with st.expander(f"🔍 Detailed Review: {sel_user['name']}", expanded=True):
            dc1, dc2, dc3 = st.columns(3)

            with dc1:
                st.markdown("**Assigned Roles**")
                for role in sel_user["assigned_roles"]:
                    info = ROLE_CATALOG.get(role, {})
                    sens = info.get("sensitivity", "—")
                    label = info.get("label", role)
                    st.markdown(
                        f"• {label} {sensitivity_badge(sens)}",
                        unsafe_allow_html=True,
                    )

            with dc2:
                st.markdown("**Excess Roles** 🔴")
                if sel_user["excess_roles"]:
                    for r in sel_user["excess_roles"]:
                        label = ROLE_CATALOG.get(r, {}).get("label", r)
                        st.markdown(f"<span style='color:#ff4444;'>⚠ {label}</span>", unsafe_allow_html=True)
                else:
                    st.success("No excess roles")

            with dc3:
                st.markdown("**SoD Conflicts** 🟠")
                if sel_user["sod_conflicts"]:
                    for c in sel_user["sod_conflicts"]:
                        st.markdown(
                            f"<span style='color:#ff8c00;'>⛔ {c['conflict_name']} ({c['severity']})</span>",
                            unsafe_allow_html=True,
                        )
                else:
                    st.success("No SoD conflicts")

            st.markdown("---")
            st.markdown("**Risk Flags**")
            for flag in sel_user["risk_flags"]:
                if "Critical" in flag or "Terminated" in flag:
                    st.error(f"🚨 {flag}")
                else:
                    st.warning(f"⚠️ {flag}")

            # Recommendation
            st.markdown("---")
            st.markdown("**📋 Recommendation**")
            if sel_user["risk_level"] == "Critical":
                st.error("IMMEDIATE ACTION: Conduct emergency access review. Revoke excess roles and resolve all SoD conflicts within 24 hours. Escalate to CISO.")
            elif sel_user["risk_level"] == "High":
                st.warning("HIGH PRIORITY: Schedule access review within 7 days. Address SoD conflicts and enforce MFA. Document risk acceptance for any retained excess access.")
            elif sel_user["risk_level"] == "Medium":
                st.info("REVIEW REQUIRED: Include in next quarterly access certification. Verify business justification for current role assignments.")
            else:
                st.success("LOW RISK: User access is within expected parameters. Continue routine monitoring.")


# ═══════════════════════════════════════════════════════════════════
# TAB 3 — Role Mapping (RBAC)
# ═══════════════════════════════════════════════════════════════════
with tab3:
    st.markdown("### 🗂 Role-Based Access Control (RBAC) Mapping")

    rbac_rows = []
    for u in analyzed_users:
        expected = RBAC_MATRIX.get(u["job_title"], [])
        excess = u["excess_roles"]
        compliant = len(excess) == 0 and not u["missing_roles"]
        rbac_rows.append({
            "User": u["name"],
            "Job Title": u["job_title"],
            "Expected Roles": ", ".join(expected) if expected else "—",
            "Assigned Roles": ", ".join(u["assigned_roles"]),
            "Excess Roles": ", ".join(excess) if excess else "—",
            "RBAC Compliant": "✅" if compliant else "❌",
        })

    rbac_df = pd.DataFrame(rbac_rows)

    def color_compliance(val):
        if val == "✅":
            return "background-color: rgba(57,255,20,0.15); color: #39ff14;"
        else:
            return "background-color: rgba(255,51,51,0.15); color: #ff4444;"

    styled_rbac = rbac_df.style.map(color_compliance, subset=["RBAC Compliant"])
    st.dataframe(styled_rbac, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── Role sensitivity heatmap ──────────────────────────────────
    st.markdown("#### 🌡 Role Sensitivity Matrix")
    sens_levels = ["Low", "Medium", "High", "Critical"]
    roles_sorted = sorted(ROLE_CATALOG.keys())
    z_matrix = []
    for role in roles_sorted:
        sens = ROLE_CATALOG[role]["sensitivity"]
        row = [1 if s == sens else 0 for s in sens_levels]
        z_matrix.append(row)

    fig_hm = go.Figure(data=go.Heatmap(
        z=z_matrix,
        x=sens_levels,
        y=[ROLE_CATALOG[r]["label"] for r in roles_sorted],
        colorscale=[[0, "#161b22"], [1, "#ff3333"]],
        showscale=False,
        hovertemplate="Role: %{y}<br>Sensitivity: %{x}<extra></extra>",
    ))
    dark_plotly_layout(fig_hm, "Role Sensitivity Heatmap")
    fig_hm.update_layout(height=480, yaxis=dict(autorange="reversed"))
    st.plotly_chart(fig_hm, use_container_width=True)

    # ── Role catalog table ────────────────────────────────────────
    st.markdown("#### 📚 Role Catalog")
    catalog_df = pd.DataFrame([
        {
            "Role Name": v["label"],
            "Role Key": k,
            "Category": v["category"],
            "Sensitivity": v["sensitivity"],
        }
        for k, v in ROLE_CATALOG.items()
    ]).sort_values("Category")

    def color_sensitivity(val):
        colors = {"Critical": "#ff3333", "High": "#ff8c00", "Medium": "#ffd700", "Low": "#39ff14"}
        c = colors.get(val, "#8b949e")
        return f"color: {c}; font-weight: 700;"

    styled_catalog = catalog_df.style.map(color_sensitivity, subset=["Sensitivity"])
    st.dataframe(styled_catalog, use_container_width=True, hide_index=True)


# ═══════════════════════════════════════════════════════════════════
# TAB 4 — SoD Conflict Analysis
# ═══════════════════════════════════════════════════════════════════
with tab4:
    st.markdown("### ⛔ Segregation of Duties (SoD) Conflict Analysis")

    # ── SoD rules reference ───────────────────────────────────────
    st.markdown("#### 📜 SoD Rules Reference")
    rules_md = ""
    for rule in SOD_RULES:
        sev = rule["severity"]
        badge = f'<span class="badge badge-{sev.lower()}">{sev}</span>'
        rules_md += f"| `{rule['role_a']}` | `{rule['role_b']}` | {rule['conflict_name']} | {badge} |\n"

    st.markdown(f"""
    | Role A | Role B | Conflict Description | Severity |
    |--------|--------|---------------------|----------|
    {rules_md}
    """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Metrics ───────────────────────────────────────────────────
    total_violations = len(sod_df)
    crit_violations = len(sod_df[sod_df["Severity"] == "Critical"]) if not sod_df.empty else 0
    high_violations = len(sod_df[sod_df["Severity"] == "High"]) if not sod_df.empty else 0
    users_affected = sod_df["User"].nunique() if not sod_df.empty else 0

    mc1, mc2, mc3, mc4 = st.columns(4)
    mc1.metric("Total Violations", total_violations)
    mc2.metric("🔴 Critical", crit_violations)
    mc3.metric("🟠 High", high_violations)
    mc4.metric("👥 Users Affected", users_affected)

    st.markdown("---")

    # ── Violations table ──────────────────────────────────────────
    if sod_df.empty:
        st.success("No SoD violations detected.")
    else:
        def color_sod_severity(val):
            colors = {"Critical": "#ff3333", "High": "#ff8c00", "Medium": "#ffd700"}
            c = colors.get(val, "#8b949e")
            return f"color: {c}; font-weight: 700;"

        styled_sod = sod_df.style.map(color_sod_severity, subset=["Severity"])
        st.dataframe(styled_sod, use_container_width=True, hide_index=True)

        st.markdown("---")

        col_a, col_b = st.columns(2)

        # Bar: violations by severity
        with col_a:
            sev_counts = sod_df["Severity"].value_counts().reset_index()
            sev_counts.columns = ["Severity", "Count"]
            fig_sev = px.bar(
                sev_counts, x="Severity", y="Count",
                color="Severity",
                color_discrete_map={"Critical": "#ff3333", "High": "#ff8c00", "Medium": "#ffd700"},
                text="Count",
            )
            fig_sev.update_traces(textposition="outside")
            dark_plotly_layout(fig_sev, "SoD Violations by Severity")
            st.plotly_chart(fig_sev, use_container_width=True)

        # Bar: violations by department
        with col_b:
            dept_counts = sod_df["Department"].value_counts().reset_index()
            dept_counts.columns = ["Department", "Violations"]
            fig_dept_sod = px.bar(
                dept_counts, x="Department", y="Violations",
                color="Violations",
                color_continuous_scale=["#58a6ff", "#ff3333"],
                text="Violations",
            )
            fig_dept_sod.update_traces(textposition="outside")
            dark_plotly_layout(fig_dept_sod, "SoD Violations by Department")
            st.plotly_chart(fig_dept_sod, use_container_width=True)


# ═══════════════════════════════════════════════════════════════════
# TAB 5 — Recommendations
# ═══════════════════════════════════════════════════════════════════
with tab5:
    st.markdown("### ✅ Audit Recommendations")

    # Metrics
    total_recs = len(recommendations)
    crit_recs = sum(1 for r in recommendations if r["priority"] == "Critical")
    high_recs = sum(1 for r in recommendations if r["priority"] == "High")
    timelines = []
    for r in recommendations:
        if "24" in r["timeline"] or "Immediate" in r["timeline"]:
            timelines.append(1)
        elif "7 days" in r["timeline"]:
            timelines.append(7)
        elif "14 days" in r["timeline"]:
            timelines.append(14)
        elif "30 days" in r["timeline"]:
            timelines.append(30)
        elif "60 days" in r["timeline"]:
            timelines.append(60)
        else:
            timelines.append(30)
    avg_timeline = round(sum(timelines) / len(timelines)) if timelines else 0

    rc1, rc2, rc3, rc4 = st.columns(4)
    rc1.metric("Total Recommendations", total_recs)
    rc2.metric("🔴 Critical Priority", crit_recs)
    rc3.metric("🟠 High Priority", high_recs)
    rc4.metric("⏱ Avg Timeline", f"{avg_timeline} days")

    st.markdown("---")

    # ── Recommendation cards ──────────────────────────────────────
    for rec in recommendations:
        priority = rec["priority"]
        p_color = {"Critical": "#ff3333", "High": "#ff8c00", "Medium": "#ffd700"}.get(priority, "#39ff14")
        border_col = {"Critical": "#ff3333", "High": "#ff8c00", "Medium": "#ffd700"}.get(priority, "#30363d")

        with st.expander(f"{rec['id']}  —  {rec['finding'][:80]}{'...' if len(rec['finding']) > 80 else ''}", expanded=False):
            st.markdown(f"""
            <div style="
                background: linear-gradient(135deg, #161b22, #1c2333);
                border-left: 4px solid {border_col};
                border-radius: 8px;
                padding: 18px 22px;
            ">
                <div style="display:flex; gap:16px; align-items:center; margin-bottom:12px;">
                    <span style="background:{p_color}33; color:{p_color}; padding:3px 12px; border-radius:20px;
                                 font-size:0.78rem; font-weight:700; border:1px solid {p_color};">{priority}</span>
                    <span style="color:#58a6ff; font-weight:700;">{rec['id']}</span>
                    <span style="color:#8b949e; font-size:0.85rem;">⏱ {rec['timeline']}</span>
                </div>
                <p style="color:#c9d1d9; margin-bottom:10px;"><strong>Finding:</strong> {rec['finding']}</p>
                <p style="color:#adbac7; margin-bottom:10px;"><strong>Recommended Action:</strong> {rec['action']}</p>
                <p style="color:#8b949e; margin:0;"><strong>Risk Owner:</strong> {rec['owner']}</p>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Export recommendations to Excel ───────────────────────────
    def build_rec_excel():
        wb = Workbook()
        ws = wb.active
        ws.title = "Recommendations"
        rec_df = pd.DataFrame([
            {
                "Rec ID": r["id"], "Priority": r["priority"],
                "Finding": r["finding"], "Action": r["action"],
                "Owner": r["owner"], "Timeline": r["timeline"],
                "Status": "Open",
            }
            for r in recommendations
        ])
        for r_idx, row in enumerate(
            [rec_df.columns.tolist()] + rec_df.values.tolist(), start=1
        ):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        style_ws(ws, rec_df)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    st.download_button(
        label="📥 Export Recommendations to Excel",
        data=build_rec_excel(),
        file_name="IAM_Recommendations_FY2025.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
